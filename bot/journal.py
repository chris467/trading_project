import openpyxl
from openpyxl import load_workbook
import pandas as pd
import os
from datetime import datetime

# Append trade record to Excel log
def append_trade_excel(record, filename="logs/trade_log.xlsx"):
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = list(record.keys())
        ws.append(headers)

    ws.append(list(record.values()))
    wb.save(filename)

# define schema
LOG_COLUMNS = [
    "ts_decision_utc", "symbol", "timeframe", "bar_ts", "signal", "prob_up",
    "entry_rule", "price_ref", "position_before", "order_side", "order_qty",
    "order_type", "order_id", "order_status", "fill_price", "fees",
    "position_after", "sl", "tp", "reason", "features_snapshot",
    "model_version", "pnl_realized", "pnl_unrealized"
]

# Append trade record to daily CSV log
def append_trade_csv(record: dict):
    date_str = datetime.utcnow().strftime("%Y%m%d")
    path = f"logs/trade_log_{date_str}.csv"

    if not os.path.exists(path):
        pd.DataFrame(columns=LOG_COLUMNS).to_csv(path, index=False)

    df = pd.read_csv(path)
    df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
    df.to_csv(path, index=False)

# Example usage
dummy = {col: "test" for col in LOG_COLUMNS}
dummy["ts_decision_utc"] = datetime.utcnow().isoformat()
append_trade_csv(dummy)